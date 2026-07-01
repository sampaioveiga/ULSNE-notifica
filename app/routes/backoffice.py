from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required
from flask_wtf import FlaskForm
from wtforms import TextAreaField, SelectField, SubmitField
from wtforms.validators import DataRequired, Length, Optional
from ..models import db, Notification, Comment, STATUS_ABERTA, STATUS_EM_ANALISE, STATUS_ENCERRADO, STATUS_LABELS
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

bp = Blueprint('backoffice', __name__, url_prefix='/admin')


class CommentForm(FlaskForm):
    text = TextAreaField('Comentário', validators=[DataRequired(), Length(min=3, max=2000)])
    submit = SubmitField('Adicionar Comentário')


class StatusForm(FlaskForm):
    status = SelectField('Mudar Estado', choices=[
        (STATUS_EM_ANALISE, STATUS_LABELS[STATUS_EM_ANALISE]),
        (STATUS_ENCERRADO, STATUS_LABELS[STATUS_ENCERRADO]),
    ])
    submit = SubmitField('Atualizar Estado')


@bp.route('/notificacoes')
@login_required
def notifications():
    status_filter = request.args.get('status', '')
    sort = request.args.get('sort', 'created_at')
    order = request.args.get('order', 'desc')

    query = Notification.query
    if status_filter:
        query = query.filter_by(status=status_filter)

    col = getattr(Notification, sort, Notification.created_at)
    if order == 'asc':
        query = query.order_by(col.asc())
    else:
        query = query.order_by(col.desc())

    notifications = query.all()
    return render_template(
        'admin/notifications.html',
        notifications=notifications,
        status_filter=status_filter,
        sort=sort,
        order=order,
        status_labels=STATUS_LABELS,
    )


@bp.route('/notificacoes/<int:notif_id>', methods=['GET', 'POST'])
@login_required
def notification_detail(notif_id):
    notif = Notification.query.get_or_404(notif_id)
    comment_form = CommentForm(prefix='comment')
    status_form = StatusForm(prefix='status')

    if request.method == 'POST':
        if 'comment-submit' in request.form and comment_form.validate():
            from flask_login import current_user
            comment = Comment(
                notification_id=notif.id,
                author_type='manager',
                user_id=current_user.id,
                text=comment_form.text.data,
            )
            db.session.add(comment)
            db.session.commit()
            flash('Comentário adicionado.', 'success')
            return redirect(url_for('backoffice.notification_detail', notif_id=notif_id))

        if 'status-submit' in request.form and status_form.validate():
            new_status = status_form.status.data
            if notif.status != STATUS_ENCERRADO:
                notif.status = new_status
                db.session.commit()
                flash(f'Estado atualizado para "{STATUS_LABELS[new_status]}".', 'success')
            else:
                flash('Não é possível alterar o estado de uma notificação encerrada.', 'warning')
            return redirect(url_for('backoffice.notification_detail', notif_id=notif_id))

    return render_template(
        'admin/detail.html',
        notif=notif,
        comment_form=comment_form,
        status_form=status_form,
        STATUS_EM_ANALISE=STATUS_EM_ANALISE,
        STATUS_ENCERRADO=STATUS_ENCERRADO,
    )


@bp.route('/notificacoes/export/excel')
@login_required
def export_excel():
    status_filter = request.args.get('status', '')

    query = Notification.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    notifications = query.order_by(Notification.created_at.desc()).all()

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Notificações'

    # Estilos
    header_fill = PatternFill(start_color='096f35', end_color='096f35', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = [
        'ID', 'Token', 'Data Incidente', 'Unidade', 'Local Incidente',
        'Tipo(s) Incidente', 'Vítima', 'Agressor', 'Estado',
        'Data Submissão', 'Danos Físicos', 'Impacto Psicológico'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Dados
    for row_idx, notif in enumerate(notifications, 2):
        row_data = [
            notif.id,
            notif.token[:12] + '...',
            notif.data_incidente.strftime('%d/%m/%Y %H:%M'),
            notif.unidade_local,
            notif.local_incidente,
            ', '.join(notif.tipos_incidente_labels),
            {'profissional': 'Profissional', 'utente': 'Utente', 'familiar': 'Familiar', 'outro': 'Outro'}.get(notif.tipo_vitima, notif.tipo_vitima),
            {'utente': 'Utente', 'familiar': 'Familiar', 'profissional': 'Profissional', 'externa': 'Externa', 'desconhecido': 'Desconhecido'}.get(notif.tipo_agressor, notif.tipo_agressor),
            notif.status_label,
            notif.created_at.strftime('%d/%m/%Y %H:%M'),
            'Sim' if notif.danos_fisicos == 'sim' else 'Não' if notif.danos_fisicos == 'nao' else 'Desconhecido',
            'Sim' if notif.impacto_psicologico == 'sim' else 'Não' if notif.impacto_psicologico == 'nao' else 'Desconhecido',
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Ajustar larguras
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 18

    # Altura das linhas
    ws.row_dimensions[1].height = 25
    for row in range(2, len(notifications) + 2):
        ws.row_dimensions[row].height = 30

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'notificacoes_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
